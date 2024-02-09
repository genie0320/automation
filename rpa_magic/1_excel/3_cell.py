from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'GeniesSheet'

ws['A1'] = 'text'
ws['A2'] = 'text_03'
ws['A3'] = 'text_100'
ws['B1'] = 'text_200'
ws['B2'] = 'text_400'
ws['B3'] = 'text_700'

print(ws['A1'].value) # bring a cell value.
print(ws['A10'].value) # return None if there is no data.

print(ws.cell(row=1, column=1).value) # bring A1 value.
print(ws.cell(row=1, column=2).value) # bring B1 value.

c= ws.cell(column =3, row=1, value = 10) # 이렇게 value를 직접 넣어줄 수도 있다.
print(c.value)

from random import *
index = 1
for x in range(1,11):
    for y in range(1, 11):
        # ws.cell(row = x, column= y, value=randint(0,100))
        ws.cell(row = x, column= y, value=index)
        index +=1
wb.save('sample.xlsx')
wb.close()
