from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()
ws.title = 'MySheet'
ws.sheet_properties.tabColor = 'ff66ff'

ws1 = wb.create_sheet('Yoursheet')
ws2 = wb.create_sheet('newsheet',2) #2번쨰 인덱스에 시트 생성

new_ws = wb['newsheet'] # 딕셔너리 형태로 시트에 접근가능.

print(wb.sheetnames) # confirm all sheets name.

new_ws['A1'] = 'text'
target = wb.copy_worksheet(new_ws) # copy to a sheets
target.title = 'Copyed sheet'

wb.save('sample.xlsx')
wb.close()
