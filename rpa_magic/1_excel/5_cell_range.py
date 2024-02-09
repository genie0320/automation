from openpyxl import Workbook, load_workbook
from random import *
wb = Workbook()
ws = wb.active
ws.append(['번호', '영어', '수학'])
for i in range(1,11):
    ws.append([i,randint(0,100), randint(0,100)])

# col_B = ws['B'] # 한줄 데이터만 가지고 오기.
# for cell in col_B:
#     print(cell.value, end = ' ')

# col_range = ws['B:C'] # B,C 데이터만 가지고 오기.
# for cols in col_range:
#     for cell in cols:
#         print(cell.value, end = ' ')
#     print('\n')

# row_title = ws[1] # 1번쨰 줄만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 일반적인 슬라이스와 다르다. 한개 빼는게 아니라 그대로 가지고 온다. 
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = ' ')
#     print()

from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end = ' ')
#         # print(cell.coordinate, end=' ')
#         xy = coordinate_from_string(cell.coordinate) # 여기 아래로는, xy 축을 나눠서 표시할 수 있게 해주는 기능.
#         print(xy[0], end='')
#         print(xy[1], end=' ')
#     print()

    # A2 B2 C2 
    # A3 B3 C3 
    # A4 B4 C4
    # A5 B5 C5
    # A6 B6 C6
    # A7 B7 C7
    # A8 B8 C8
    # A9 B9 C9
    # A10 B10 C10
    # A11 B11 C11

# 한 행 또는 한 열만 가져오기.
# print(tuple(ws.rows)) # 한 열전체 가져오기.
# print(tuple(ws.columns)) # 한 줄전체 가져오기

# for row in tuple(ws.rows):
#     print(row[2].value) 

# for column in tuple(ws.columns):
#     print(column[2].value)

# iter를 사용했을 때 좋은 점은, 파라메터를 전달해서 원하는 값을 가져올 수 있다.
# for row in ws.iter_rows(): # 위와 비슷한 형식으로 노출됨.
#     print(row[1].value)
#     print(row)

# for column in ws.iter_cols(): # 위와 비슷한 형식으로 노출됨.
#     print(column[1].value)
#     print(column)

# 아래처럼 하면, 한 '행'을 중심으로 가져오고. col로 바꾸면 열로 가져온다.
# 굳이 min, max를 다 정해줄 필요는 없고, 안 정해주면 한계까지 가져온다.
for row in ws.iter_rows(min_row=1, max_row =5, min_col = 2, max_col = 5): # 위와 비슷한 형식으로 노출됨.
    print(row[0].value, row[1].value)
    print(row)
   
wb.save('sample02.xlsx')
