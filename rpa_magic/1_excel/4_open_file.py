from openpyxl import Workbook, load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active

# for x in range(1, 11):
#     for y in range(1,11):
#         print(ws.cell(row=x, column=y).value, end= " ")
#     print()

for x in range(1, ws.max_row +1): # 명 칸이 있는지 모를때는 max_row등을 쓸 수 있다.
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end= " ")
    print()

